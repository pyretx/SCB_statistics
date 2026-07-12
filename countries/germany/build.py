"""Destatis Verdiensterhebung table 62361-03 → bundled Germany snapshot.

Statistisches Bundesamt (Destatis) publishes gross MONTHLY earnings by occupation
(KldB 2010) as a no-auth downloadable statistical report ("Verdienste nach
Branchen und Berufen"). The GENESIS REST API needs an account, but this report
XLSX is public — so, like the UK/US builds, we parse it offline and ship a small
gzipped snapshot the provider reads.

Table 62361-03 "Bruttomonatsverdienste (ohne Sonderzahlungen) nach beruflichen
Tätigkeiten (KldB 2010), Vollzeitbeschäftigte" has three CSV sub-sheets by KldB
level — _1 Berufshauptgruppen (2-digit), _2 Berufsgruppen (3-digit), _3
Berufsgattungen (5-digit). Each row carries the KldB code ("KB10-11101"), the
German occupation name, and two measures: VST047 = mean (arithmetisches Mittel)
and VST052 = median (Zentralwert), both gross monthly EUR excluding special
payments, for full-time employees. The reference is the April survey month.

KldB nests by prefix (5-digit 11101 → 3-digit 111 → 2-digit 11), so the shared
drill-down / code browser work; the 4-digit level is simply not published here.
"""
from __future__ import annotations

import datetime
import gzip
import io
import json
import os
import tempfile

import requests

_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUT = os.path.join(_ROOT, "germany_kldb.json.gz")
_APP_SETTINGS = os.path.join(_ROOT, "app_settings.json")
_CACHE = os.path.join(tempfile.gettempdir(), "destatis_build_cache")
_UA = {"User-Agent": "Mozilla/5.0 (salary-explorer; research)"}

# Public Destatis statistical-report download ("Verdienste nach Branchen und
# Berufen"). The hashed filename changes when Destatis revises the report; the
# admin refresh can re-resolve it from the downloads page later.
REPORT_URL = ("https://www.destatis.de/DE/Themen/Arbeit/Verdienste/"
              "Verdienste-Branche-Berufe/Publikationen/Downloads/"
              "statistischer-bericht-verdienste-5623601253245.xlsx?__blob=publicationFile&v=2")
_SHEETS = {"csv-62361-03_1": 2, "csv-62361-03_2": 3, "csv-62361-03_3": 5}
_CODE_COL, _NAME_COL, _MEAN_COL, _MEDIAN_COL = 11, 12, 13, 14
DEFAULT_LATEST_YEAR = 2025
STAT_COLS = ["mean", "median"]


def latest_year() -> int:
    try:
        with open(_APP_SETTINGS, encoding="utf-8") as f:
            return int(json.load(f).get("germany_latest_year", DEFAULT_LATEST_YEAR))
    except Exception:
        return DEFAULT_LATEST_YEAR


def save_latest_year(year: int):
    data = {}
    try:
        with open(_APP_SETTINGS, encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        pass
    data["germany_latest_year"] = int(year)
    with open(_APP_SETTINGS, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _download(url: str = REPORT_URL) -> bytes:
    os.makedirs(_CACHE, exist_ok=True)
    fp = os.path.join(_CACHE, "verdienste_berufe.xlsx")
    if os.path.exists(fp):
        return open(fp, "rb").read()
    r = requests.get(url, headers=_UA, timeout=300)
    r.raise_for_status()
    open(fp, "wb").write(r.content)
    return r.content


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(round(v))
    s = str(v).strip().replace(".", "").replace(",", ".")
    if s in ("", "/", "-", "x", ".", "…", "()"):
        return None
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def build(url: str = REPORT_URL, out_path: str = OUT, log=print) -> dict:
    import openpyxl
    log("downloading Destatis Verdienste-nach-Berufen report …")
    wb = openpyxl.load_workbook(io.BytesIO(_download(url)), read_only=True, data_only=True)
    codes: dict[str, str] = {}
    stats: dict[str, list] = {}
    year = DEFAULT_LATEST_YEAR
    for sheet, exp_len in _SHEETS.items():
        if sheet not in wb.sheetnames:
            log(f"  WARNING: sheet {sheet} missing")
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        added = 0
        for r in rows[1:]:
            if len(r) <= _MEDIAN_COL:
                continue
            raw = r[_CODE_COL]
            if not raw:
                continue
            code = str(raw).strip().replace("KB10-", "")
            if not code.isdigit() or len(code) != exp_len:
                continue
            name = str(r[_NAME_COL]).strip() if r[_NAME_COL] else code
            codes[code] = name
            stats[code] = [_num(r[_MEAN_COL]), _num(r[_MEDIAN_COL])]
            added += 1
            # capture reference year from the Zeit column (col 4)
            zt = r[4]
            if hasattr(zt, "year"):
                year = zt.year
            elif isinstance(zt, str) and zt[:4].isdigit():
                year = int(zt[:4])
        log(f"  {sheet}: {added} codes (len {exp_len})")

    leaves = sum(1 for c in codes if len(c) == 5)
    payload = {
        "built_at": datetime.date.today().isoformat(),
        "source": url.split("?")[0],
        "classification": "KldB 2010 (Destatis Verdiensterhebung 62361-03)",
        "note": "Gross MONTHLY earnings (EUR, excl. special payments), full-time; "
                "mean (VST047) + median (VST052). Names are official German KldB titles.",
        "year": year,
        "currency": "EUR",
        "stat_cols": STAT_COLS,
        "codes": {"DE": codes},
        "stats": stats,
    }
    tmp = out_path + ".tmp"
    with gzip.open(tmp, "wt", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, out_path)
    size = os.path.getsize(out_path)
    log(f"wrote {len(codes)} KldB codes ({leaves} leaf), year {year} ({size / 1e6:.2f} MB)")
    return {"built_at": payload["built_at"], "year": year, "codes": len(codes),
            "leaves": leaves, "size": size}


def bundled_info(path: str = OUT) -> dict:
    try:
        with gzip.open(path, "rt", encoding="utf-8") as f:
            d = json.load(f)
        return {"built_at": d.get("built_at"), "year": d.get("year"),
                "size": os.path.getsize(path), "source": d.get("source")}
    except Exception:
        return {}


if __name__ == "__main__":
    print(build())
