"""ONS ASHE Table 14 (occupation, 4-digit SOC 2020) → bundled UK snapshot.

ONS publishes occupation earnings only as spreadsheets (not via Nomis/API), so —
like the US OEWS build — we parse the official workbooks offline and ship a small
gzipped snapshot the provider reads. This module SHIPS with the app so the admin
panel can trigger a refresh at runtime; openpyxl is imported lazily.

Source: "Earnings and hours worked, occupation by four-digit SOC: ASHE Table 14",
table 14.7a "Annual pay - Gross". ASHE moved to SOC 2020 from the 2021 data year,
so we bundle 2021-2024 (revised) for a consistent-classification trend.

Each workbook has sheets All / Male / Female (→ total / men / women). Rows carry
the SOC hierarchy in the Code column (1 / 11 / 111 / 1111 = major / sub-major /
minor / unit group) and the name in Description. Measures: number of jobs
(thousands), median, mean, and percentiles 10/20/25/30/40/60/70/75/80/90 — we
keep P10·P25·median·P75·P90 (the model's columns). '..'/'x' = suppressed → null.

Values are stored as a GROSS MONTHLY figure = official gross ANNUAL pay ÷ 12, so
the UK is comparable with the app's other (monthly) countries; this is an exact
transform of an official figure, not a modelled estimate.
"""
from __future__ import annotations

import datetime
import gzip
import io
import json
import os
import tempfile
import zipfile

import requests

_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUT = os.path.join(_ROOT, "uk_ashe.json.gz")
_APP_SETTINGS = os.path.join(_ROOT, "app_settings.json")
_CACHE = os.path.join(tempfile.gettempdir(), "ashe_build_cache")
_UA = {"User-Agent": "Mozilla/5.0 (salary-explorer; research)"}

BASE = ("https://www.ons.gov.uk/file?uri=/employmentandlabourmarket/peopleinwork/"
        "earningsandworkinghours/datasets/occupation4digitsoc2010ashetable14/")
# reference year → zip path (revised = final; SOC 2020 basis from 2021 on)
_YEAR_ZIP = {
    2021: "2021revised/ashetable142021revised.zip",
    2022: "2022revised/ashetable142022revised.zip",
    2023: "2023revised/ashetable142023revised.zip",
    2024: "2024revised/ashetable142024revised.zip",
}
DEFAULT_LATEST_YEAR = 2024
_SEX_SHEET = {"total": "All", "men": "Male", "women": "Female"}
STAT_COLS = ["count", "median", "mean", "p10", "p25", "p75", "p90"]


def latest_year() -> int:
    try:
        with open(_APP_SETTINGS, encoding="utf-8") as f:
            return int(json.load(f).get("uk_latest_year", DEFAULT_LATEST_YEAR))
    except Exception:
        return DEFAULT_LATEST_YEAR


def save_latest_year(year: int):
    data = {}
    try:
        with open(_APP_SETTINGS, encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        pass
    data["uk_latest_year"] = int(year)
    with open(_APP_SETTINGS, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── helpers ──────────────────────────────────────────────────────────────────
def _download(url: str) -> bytes:
    os.makedirs(_CACHE, exist_ok=True)
    fp = os.path.join(_CACHE, url.rsplit("/", 1)[1].split("?")[0])
    if os.path.exists(fp):
        return open(fp, "rb").read()
    r = requests.get(url, headers=_UA, timeout=600)
    r.raise_for_status()
    open(fp, "wb").write(r.content)
    return r.content


def _num_month(v):
    """Annual money value → rounded MONTHLY int (annual ÷ 12); None if suppressed."""
    n = _num_raw(v)
    return None if n is None else int(round(n / 12))


def _num_raw(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "x", "..", ".", "-", ":", "#", "*"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _read_workbook(raw: bytes, member: str) -> dict:
    """{sheet_name: [rows]} for either .xlsx (openpyxl) or legacy .xls (xlrd).
    Rows are plain lists so the parser is format-agnostic."""
    if member.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        return {sh: [list(r) for r in wb[sh].iter_rows(values_only=True)]
                for sh in wb.sheetnames}
    import xlrd
    wb = xlrd.open_workbook(file_contents=raw)
    return {sh.name: [sh.row_values(i) for i in range(sh.nrows)]
            for sh in wb.sheets()}


def _find_header(rows):
    """Locate the header row (has 'Description' + 'Code') and map the columns we
    need. Returns (header_index, colmap). Robust to column drift across years."""
    for ri, row in enumerate(rows[:12]):
        cells = [("" if c is None else str(c).strip()) for c in row]
        if "Description" in cells and "Code" in cells:
            cm = {}
            for j, c in enumerate(cells):
                if c == "Description":
                    cm["desc"] = j
                elif c == "Code":
                    cm["code"] = j
                elif c == "(thousand)":
                    cm["count"] = j
                elif c == "Median":
                    cm["median"] = j
                elif c == "Mean":
                    cm["mean"] = j
                elif c in ("10", "25", "75", "90"):
                    cm[f"p{c}"] = j
            return ri, cm
    return None, {}


def _parse_sheet(rows, monthly=True):
    """→ ({code: [count, median, mean, p10, p25, p75, p90]}, {code: name})."""
    hi, cm = _find_header(rows)
    if hi is None or "code" not in cm or "median" not in cm:
        return {}, {}
    out, names = {}, {}
    money = ("median", "mean", "p10", "p25", "p75", "p90")
    for row in rows[hi + 1:]:
        raw_code = row[cm["code"]] if cm["code"] < len(row) else None
        if raw_code is None or raw_code == "":
            continue
        code = str(raw_code).strip()
        if code.endswith(".0"):                     # xlrd reads codes as floats
            code = code[:-2]
        if not code.isdigit() or not (1 <= len(code) <= 4):
            continue
        desc = row[cm["desc"]] if cm["desc"] < len(row) else None
        names[code] = str(desc).strip() if desc else code
        vals = []
        for key in STAT_COLS:
            col = cm.get(key)
            cell = row[col] if (col is not None and col < len(row)) else None
            if key == "count":
                n = _num_raw(cell)
                vals.append(None if n is None else int(round(n * 1000)))
            elif key in money and monthly:
                vals.append(_num_month(cell))
            else:
                n = _num_raw(cell)
                vals.append(None if n is None else int(round(n)))
        out[code] = vals
    return out, names


def _fetch_cpi(years) -> dict:
    """ONS CPIH all-items annual index (series L522, dataset MM23, 2015=100) →
    {year: index} for the trend tab's real-terms overlay. CPIH is ONS's preferred
    inflation measure. Best-effort — empty on any failure."""
    try:
        url = "https://www.ons.gov.uk/economy/inflationandpriceindices/timeseries/l522/mm23/data"
        r = requests.get(url, headers=_UA, timeout=60)
        r.raise_for_status()
        out = {}
        for rec in r.json().get("years", []):
            y = int(rec["date"])
            if y in years:
                out[str(y)] = float(rec["value"])
        return out
    except Exception:
        return {}


# ── build ────────────────────────────────────────────────────────────────────
def build(years=None, out_path: str = OUT, log=print) -> dict:
    years = sorted(years or _YEAR_ZIP.keys())
    stats: dict[str, dict] = {}
    codes: dict[str, str] = {}
    latest = max(years)
    for yr in years:
        url = BASE + _YEAR_ZIP[yr]
        log(f"downloading ASHE {yr} … {_YEAR_ZIP[yr].rsplit('/', 1)[1]}")
        z = zipfile.ZipFile(io.BytesIO(_download(url)))
        member = [m for m in z.namelist()
                  if "14.7a" in m and m.lower().endswith((".xls", ".xlsx"))][0]
        sheets = _read_workbook(z.read(member), member)
        ystats = {}
        for sex, sheet in _SEX_SHEET.items():
            if sheet not in sheets:
                continue
            parsed, names = _parse_sheet(sheets[sheet])
            ystats[sex] = parsed
            if yr == latest and sex == "total":
                codes = names
        stats[str(yr)] = ystats
        log(f"  {yr}: " + ", ".join(f"{s}={len(v)}" for s, v in ystats.items()))

    cpi = _fetch_cpi(set(years))
    leaves = sum(1 for c in codes if len(c) == 4)
    payload = {
        "built_at": datetime.date.today().isoformat(),
        "source": BASE + _YEAR_ZIP[latest],
        "classification": "SOC 2020 (ONS ASHE Table 14.7a, Annual pay - Gross)",
        "note": "Gross MONTHLY pay = official gross annual pay / 12; suppressed cells null.",
        "years": years,
        "latest_year": latest,
        "currency": "GBP",
        "stat_cols": STAT_COLS,
        "codes": {"EN": codes},
        "cpi": cpi,
        "stats": stats,
    }
    tmp = out_path + ".tmp"
    with gzip.open(tmp, "wt", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, out_path)
    size = os.path.getsize(out_path)
    log(f"wrote {len(codes)} SOC codes ({leaves} leaf), {len(years)} years, "
        f"CPI={len(cpi)} pts ({size / 1e6:.2f} MB)")
    return {"built_at": payload["built_at"], "years": years, "latest": latest,
            "codes": len(codes), "leaves": leaves, "cpi": len(cpi), "size": size}


def bundled_info(path: str = OUT) -> dict:
    try:
        with gzip.open(path, "rt", encoding="utf-8") as f:
            d = json.load(f)
        return {"built_at": d.get("built_at"), "year": d.get("latest_year"),
                "years": d.get("years"), "size": os.path.getsize(path),
                "source": d.get("source")}
    except Exception:
        return {}


if __name__ == "__main__":
    print(build())
